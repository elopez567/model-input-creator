import pyodbc
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from datetime import date

class Mainapp:
    def __init__(self):
        app = Tk()
        app.title('BestExInputCreator v1.0')
        app.geometry("300x180")
        app['background'] = '#003087'
        
        # Labels
        file_label = Label(app, bg='#003087', fg='white', text = "Showout File")
        
        space  = Label(app, bg='#003087', text = "      ")
        space2 = Label(app, bg='#003087', text = "      ")
        space3 = Label(app, bg='#003087', text = "      ")        
        space4 = Label(app, bg='#003087', text = "      ")
        space5 = Label(app, bg='#003087', text = "      ")
        space6 = Label(app, bg='#003087', text = "      ")

        
        # File Path Text Box
        filedir = StringVar()
        file_dir = Entry(app, textvariable=filedir, width=25, borderwidth=3).grid(row=8,column=3, columnspan=5, sticky=tk.W+tk.E)
        
        # Grids
        space.grid(row=0, column=0)
        space2.grid(row=9,column=0)
        space3.grid(row=11,column=0)
        file_label.grid(row=7, column=5)
        
        # Origination Date
        today = date.today()
        day = today.strftime("%m/%d/%Y")
        month = day[:2]
        month = int(month) - 1
        self.orig_date = f'{month}/1/2022'

        
        # Function to get excel worksheet names
        def get_sheetnames_xlsx(filepath):
            wb = load_workbook(filepath, read_only=True, keep_links=False)
            return wb.sheetnames
        
        # Open File Dialog Function
        def openfile():
            app.filename = filedialog.askopenfilename(initialdir=r"M:\Capital Markets\Investor Loan Analysis\Ritz", title="Select Deal Tape")
            self.sheet_names = get_sheetnames_xlsx(app.filename)
            
            filedir = StringVar()
            filedir.set(app.filename)
            file_dir = Entry(app, textvariable=filedir, width=25, borderwidth=3).grid(row=8,column=3, columnspan=5, sticky=tk.W+tk.E)
            
            options = self.sheet_names
            self.clicked = StringVar()
            self.clicked.set('Select Sheet')
            sheet_menu = OptionMenu(app, self.clicked, *options).grid(row=11, column=5)
            
        #Create Strat Function
        def runstrat():
            self.df = pd.read_excel(io=app.filename, skiprows=1, sheet_name=self.clicked.get())
            
            # Reads paths to Input file templates
            self.adco = pd.read_excel(io=r'M:\Capital Markets\Users\Emmanuel Lopez\ADCO\ADCO Templates\ADCO Input Template.xlsx')
            self.milan = pd.read_excel(io=r"M:\Capital Markets\Users\Emmanuel Lopez\Moody's MILAN\ExampleDeal.xlsx")            
            
            # ADCO tape cracking
            self.adco= self.adco.drop([0,0])
            self.adco['Loan_ID'] = self.df['Loan']
            self.adco['Original_Term'] = self.df['TERM']
            self.adco['Age'] = 0
            self.adco['Remaining_Term'] = self.df['TERM']
            self.adco['Lien_Position'] = 1
            self.adco['OrigRate'] = self.df['Rate']
            self.adco['CurrentRate'] = self.df['Rate']
            self.adco['Current_Loan_Size'] = self.df['UPB']
            self.adco['Orig_Loan_Size'] = self.df['UPB']
            self.adco['Orig_LTV'] = self.df['LTV']
            self.adco['Orig_Total_LTV'] = self.df['LTV']
            self.adco['FICO_Score'] = self.df['Fico']
            self.adco['RateFixed'] = 1
            self.adco['Balloon_Months'] = 0
            self.adco['PP_Months'] = 0
            self.adco['IO_Months'] = 0
            self.adco['Delinquency'] = 'C'
            self.adco['First_Reset_Age'] = -1
            self.adco['Months_Between_Reset'] = -1
            self.adco['Index'] = -1
            self.adco['Gross_Margin'] = -1
            self.adco['Life_Cap'] = -1
            self.adco['Life_Floor'] = -1
            self.adco['Periodic_Cap'] = -1
            self.adco['Periodic_Floor'] = -1
            self.adco['Payment_Change_Cap'] = -1
            self.adco['Payment_Reset_Period'] = -1
            self.adco['Payment_Recast_Period'] = -1
            self.adco['Current_Minimum_Payment'] = -1
            self.adco['Max_Negam_Percent'] = -1
            self.adco['Documentation'] = 'F'
            self.adco['MI_Cutoff'] = -1
            self.adco['MI_Premium'] = -1

            self.adco['Occupancy'] = self.df['Occupancy']
            self.adco['LoanPurpose'] = self.df['Purpose']
            self.adco['PropertyType'] = self.df['PropType']
            self.adco['Num_Units'] = self.df['PropType']
            self.adco['State'] = self.df['STATE'].apply(lambda x:x.strip())
            self.adco['Zip'] = self.df['Zip']
            #Deal Name 
            self.adco['Group_Number'] = 1
            self.adco['tunestr'] = 'FNMA'
            self.adco['PrevDel'] = 0
            self.adco['Months_Since_Delinq'] = -1
            self.adco['Bankrupt'] = -1
            self.adco['ServicingFee'] = -1

            self.adco['MI_Percent'] = 0
            self.adco.loc[(self.adco['Orig_LTV'] > 80) & (self.adco['Orig_LTV'] <= 85), 'MI_Percent'] = 12
            self.adco.loc[(self.adco['Orig_LTV'] > 85) & (self.adco['Orig_LTV'] <= 90), 'MI_Percent'] = 25
            self.adco.loc[(self.adco['Orig_LTV'] > 90) & (self.adco['Orig_LTV'] <= 95), 'MI_Percent'] = 30
            self.adco.loc[(self.adco['Orig_LTV'] > 95), 'MI_Percent'] = 35

            self.adco.loc[(self.adco['PropertyType']!='2 Unit') &  (self.adco['PropertyType']!='3 Unit') &  (self.adco['PropertyType']!='4 Unit'), 'Num_Units'] = 1
            self.adco.loc[(self.adco['PropertyType']=='2 Unit'), 'Num_Units'] = 2
            self.adco.loc[(self.adco['PropertyType']=='3 Unit'), 'Num_Units'] = 3
            self.adco.loc[(self.adco['PropertyType']=='4 Unit'), 'Num_Units'] = 4

            self.adco.loc[(self.adco['LoanPurpose']=='PURCH'),'LoanPurpose'] = 'P'
            self.adco.loc[(self.adco['LoanPurpose']=='REFINANCE - RATE/TERM'),'LoanPurpose'] = 'R'
            self.adco.loc[(self.adco['LoanPurpose']=='REFINANCE - CASHOUT'),'LoanPurpose'] = 'E'
            self.adco.loc[(self.adco['PropertyType']=='MFD'),'PropertyType'] = 'MH'
            self.adco.loc[(self.adco['PropertyType']=='2 Unit'),'PropertyType'] = 'MFR'
            self.adco.loc[(self.adco['PropertyType']=='3 Unit'),'PropertyType'] = 'MFR'
            self.adco.loc[(self.adco['PropertyType']=='4 Unit'),'PropertyType'] = 'MFR'
            self.adco.loc[(self.adco['Occupancy']=='OWN'),'Occupancy'] = 'O'
            self.adco.loc[(self.adco['Occupancy']=='2ND'),'Occupancy'] = 'S'
            self.adco.loc[(self.adco['Occupancy']=='NOO'),'Occupancy'] = 'I'
            
            # Moody's tape cracking
            self.milan= self.milan.drop([0,0])
            
            # Fuction to return first digit of loan number (determines channel)
            def left(var):
                return str(var)[0]

            self.milan['Loan Number'] = self.df['Loan']
            self.milan['Current Loan Amount'] = self.df['UPB']
            self.milan['Amortization Type'] = 1
            self.milan.loc[(self.adco['LoanPurpose']=='P'),'Loan Purpose'] = 7
            self.milan.loc[(self.adco['LoanPurpose']=='R'),'Loan Purpose'] = 9
            self.milan.loc[(self.adco['LoanPurpose']=='E'),'Loan Purpose'] = 3


            self.milan['Left'] = self.df['Loan'].apply(left)

            self.milan.loc[(self.milan['Left']=='7'),'Channel'] = 1
            self.milan.loc[(self.milan['Left']=='6'),'Channel'] = 2
            self.milan.loc[(self.milan['Left']=='8'),'Channel'] = 3

            self.milan['Origination Date'] = self.orig_date 
            self.milan['Original Interest Rate'] = self.df['Rate'].apply(lambda x: x/100)
            self.milan['Original Amortization Term'] = self.df['TERM']
            self.milan['Original Term to Maturity'] = self.df['TERM']
            self.milan['Original Interest Only Term'] = 0 
            self.milan['Prepayment Penalty Total Term'] = 0 
            self.milan['FICO'] = self.df['Fico']
            self.milan['Originator DTI'] = self.df['DTI'].apply(lambda x: x/100)
            self.milan['Postal Code'] = self.df['Zip']

            self.milan.loc[(self.df['PropType']=='SFR'),'Property Type'] = 1
            self.milan.loc[(self.df['PropType']=='MFD'),'Property Type'] = 6
            self.milan.loc[(self.df['PropType']=='Condo'),'Property Type'] = 3
            self.milan.loc[(self.df['PropType']=='PUD'),'Property Type'] = 6
            self.milan.loc[(self.df['PropType']=='2 Unit'),'Property Type'] = 13
            self.milan.loc[(self.df['PropType']=='3 Unit'),'Property Type'] = 14
            self.milan.loc[(self.df['PropType']=='4 Unit'),'Property Type'] = 15

            self.milan.loc[(self.adco['Occupancy']=='O'),'Occupancy'] = 1
            self.milan.loc[(self.adco['Occupancy']=='S'),'Occupancy'] = 2
            self.milan.loc[(self.adco['Occupancy']=='I'),'Occupancy'] = 3
            self.milan['Original Appraised Property Value'] = self.df['UPB'] / (self.df['LTV'].apply(lambda x: x/100))
            self.milan['Original CLTV'] = self.df['LTV'].apply(lambda x: x/100)
            self.milan['Original LTV'] = self.milan['Original CLTV']
            self.milan['LGD Model Type'] = 'PrivateLabel'
            self.milan['PD Model Type'] = 'PrivateLabel'
            self.milan['Pool ID'] = 1
            self.milan['Documentation'] = 1
            self.milan['Pre-closing Modification Flag'] = 0
            
            del self.df
            
            self.milan = self.milan.drop(labels='Left', axis=1)
            moodysSQLpull()
            FicoAdjust()
            
            # Saves finished Input files to desired output paths
            self.adco.to_excel(fr'M:\Capital Markets\Users\Emmanuel Lopez\ADCO_{self.clicked.get()}.xlsx', index=False )
            self.milan_input.to_excel(fr'M:\Capital Markets\Users\Emmanuel Lopez\Milan_{self.clicked.get()}.xlsx', index=False )
            
            del self.adco, self.milan_input
            
            # Done status display
            status = StringVar()
            status.set('Done!')
            status_entry = Entry(app, textvariable=status, width=8, bg='green', fg='white').grid(row=12,column=5, sticky=tk.W+tk.E)
            
        # Function to pull PIW & Borrower Count from SQL 
        def moodysSQLpull():
            #Open SQL Connection
            conn = pyodbc.connect('Driver={SQL Server Native Client 11.0};'
                              'Server=AWWDCORPSQLP02;'
                              'Database=clg_reporting;'
                              'Trusted_Connection=yes;')
            
            list_of_ids = list(self.milan['Loan Number'].apply(lambda x: f'{x}'))
            tup = tuple(list_of_ids)
            
            #Executing SQL queries and exporting results into dataframes
            bor_corr = pd.read_sql_query(f'''SELECT l.loannum as 'Loan Number', l.BorCount as 'Total Number of Borrowers' 
                                          FROM clg_reporting.dbo.vw_LOS_PCG_LoanDetails l
                                          WHERE l.LOS_Name = 'em' and l.TestLoanFlag = 'n'
                                          and loannum in {tup}
                                          ''', conn)
            bor_bdl = pd.read_sql_query(f'''SELECT LoanId as 'Loan Number', BorrowersCount as 'Total Number of Borrowers' 
                                            from rrd..HUB_BDL_LoanDetail
                                            where LoanId in {tup}
                                            ''', conn)
            bor_cdl = pd.read_sql_query(f'''SELECT LoanId as 'Loan Number', BorrowersCount as 'Total Number of Borrowers' 
                                            from rrd..HUB_CDL_LoanDetail
                                            where LoanId in {tup}
                                            ''', conn)
            piw = pd.read_sql_query(f'''SELECT [LoanNumber] as 'Loan Number',[property_inspection_waiver_indicator] as 'PIW' 
                                            FROM cm_inventory_management.ods_readonly.vw_pooling_data 
                                            WHERE [LoanNumber] in {tup}
                                            ''', conn)
            #Closes SQL connections 
            conn.close()
                
            #Combines BDL CDL  & Correspondent BorCount and merges into Milan input dataframe
            bor_combined = pd.concat([bor_bdl,bor_cdl,bor_corr])
            del bor_bdl, bor_cdl, bor_corr
            bor_combined["Loan Number"] = bor_combined["Loan Number"].apply(lambda x: int(x))
            bor_combined.loc[(bor_combined['Total Number of Borrowers'].isnull()), 'Total Number of Borrowers'] = 1
            bor_combined["Total Number of Borrowers"] = bor_combined["Total Number of Borrowers"].apply(lambda x: int(x))
            merge_first = pd.merge(self.milan,bor_combined, left_on="Loan Number", right_on="Loan Number", how="left")   
            del self.milan, bor_combined
            merge_first["Total Number of Borrowers_x"] = merge_first["Total Number of Borrowers_y"]
            merge_first = merge_first.drop(columns=["Total Number of Borrowers_y"])
            merge_first = merge_first.rename(columns={'Total Number of Borrowers_x':'Total Number of Borrowers'})
            merge_first.loc[(merge_first['Total Number of Borrowers'].isna()), 'Total Number of Borrowers'] = 1
            merge_first['xxx'] = ''
            piw["Loan Number"] = piw["Loan Number"].apply(lambda x: int(x))
            self.milan_input = pd.merge(merge_first,piw, on="Loan Number", how="left")
            del merge_first, piw
            self.milan_input.loc[(self.milan_input['PIW'] == True), 'PIW'] = 'Y'
            self.milan_input.loc[(self.milan_input['PIW'] != 'Y'), 'PIW'] = 'N'
            
        # Function to Adjust FICO to accurately fit ADCO model
        def FicoAdjust ():
            #DTI FICO Adjustments
            dti=self.milan_input['Originator DTI']
            self.adco.loc[(dti<=0.30), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 53.01)
            self.adco.loc[(dti>0.30) & (dti<=0.35), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 22.33)
            self.adco.loc[(dti>0.35) & (dti<=0.40), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 4.52)
            self.adco.loc[(dti>0.40) & (dti<=0.45), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 7.42)
            self.adco.loc[(dti>0.45), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 51.85)

            #Purpose FICO Adjustments
            purp = self.adco['LoanPurpose']
            self.adco.loc[(purp=='P'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 12.38)
            self.adco.loc[(purp=='E'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 68.55)
            self.adco.loc[(purp=='R'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 21.27)

            #Occupancy FICO Adjustments
            occ = self.adco['Occupancy']
            self.adco.loc[(occ=='O'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 16.99)
            self.adco.loc[(occ=='S'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 25.33)
            self.adco.loc[(occ=='I'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 88.74)

            #Property Type FICO Adjustments
            prop = self.adco['PropertyType']
            self.adco.loc[(prop=='SFR'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 3.25)
            self.adco.loc[(prop=='Condo'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 12.49)
            self.adco.loc[(prop=='MH'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 87.42)
            self.adco.loc[(prop=='PUD'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 2.07)
            self.adco.loc[(prop=='Coop'), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 91.5)

            #Borrower Count FICO Adjustments
            borCount=self.milan_input['Total Number of Borrowers']
            self.adco.loc[(borCount==1), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x - 36.49)
            self.adco.loc[(borCount==2), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 38.4)
            self.adco.loc[(borCount>2), 'FICO_Score'] = self.adco['FICO_Score'].apply(lambda x: x + 72.18)    
            
            self.adco.loc[(self.adco['FICO_Score'] > 850), 'FICO_Score'] = 850
            self.adco.loc[(self.adco['FICO_Score'] < 350), 'FICO_Score'] = 350
            
            del dti, borCount, purp, occ, prop
            
        # Buttons
        openfile_btn = Button(app, bg='#31AFDF', fg='white', activebackground='#F1C400', activeforeground='white', command=openfile,text="Open File").grid(row=10, column=4)
        runstrat_btn = Button(app, bg='#31AFDF', fg='white', activebackground='#F1C400', activeforeground='white', command=runstrat,text="Create Inputs").grid(row=10, column=6)
        
        app.mainloop()
        
Mainapp()    
    
    
